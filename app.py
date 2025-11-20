from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file
from datetime import datetime, timedelta
import json
import os
from werkzeug.security import generate_password_hash, check_password_hash
import secrets
import random

# Determine the base directory (where this script is located)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PARENT_DIR = os.path.dirname(BASE_DIR)

template_folder = os.path.join(BASE_DIR, 'templates')
app = Flask(__name__, template_folder=template_folder)

# Fixed secret key for session persistence across restarts
app.secret_key = 'weekend-trunk-shifts-secret-key-2025'

# Data storage (in production, use a proper database)
DATA_DIR = os.path.join(BASE_DIR, 'data')
os.makedirs(DATA_DIR, exist_ok=True)

EMPLOYEES_FILE = os.path.join(DATA_DIR, 'employees.json')
PREFERENCES_FILE = os.path.join(DATA_DIR, 'preferences.json')
SETTINGS_FILE = os.path.join(DATA_DIR, 'settings.json')
ASSIGNMENTS_FILE = os.path.join(DATA_DIR, 'assignments.json')

# Generate 60 weekend shifts (20 weekends starting Dec 14, 2025)
def generate_shifts():
    shifts = []
    start_date = datetime(2025, 12, 13)  # Saturday Dec 13, 2025
    shift_id = 0
    
    for week in range(20):
        saturday = start_date + timedelta(weeks=week)
        sunday = saturday + timedelta(days=1)
        
        # Saturday shift
        shifts.append({
            'id': shift_id,
            'date': saturday.strftime('%Y-%m-%d'),
            'day': 'Saturday',
            'time': '11:00 AM - 7:00 PM',
            'slots': 1,
            'week': week + 1
        })
        shift_id += 1
        
        # Sunday morning shift
        shifts.append({
            'id': shift_id,
            'date': sunday.strftime('%Y-%m-%d'),
            'day': 'Sunday',
            'time': '8:00 AM - 4:00 PM',
            'slots': 1,
            'week': week + 1
        })
        shift_id += 1
        
        # Sunday evening shift
        shifts.append({
            'id': shift_id,
            'date': sunday.strftime('%Y-%m-%d'),
            'day': 'Sunday',
            'time': '3:00 PM - 10:00 PM',
            'slots': 1,
            'week': week + 1
        })
        shift_id += 1
    
    return shifts

SHIFTS = generate_shifts()

# Initialize data files
def init_data_files():
    # Create 30 employees
    if not os.path.exists(EMPLOYEES_FILE):
        employees = {}
        
        # Manager account
        employees['admin'] = {
            'name': 'Admin',
            'is_manager': True,
            'password': generate_password_hash('admin123')
        }
        
        # 30 employee accounts
        for i in range(1, 31):
            username = f'employee{i}'
            employees[username] = {
                'name': f'Employee{i}',
                'is_manager': False,
                'password': generate_password_hash('password')
            }
        
        with open(EMPLOYEES_FILE, 'w') as f:
            json.dump(employees, f, indent=2)
    
    if not os.path.exists(PREFERENCES_FILE):
        with open(PREFERENCES_FILE, 'w') as f:
            json.dump({}, f)
    
    if not os.path.exists(SETTINGS_FILE):
        # Default deadline: 7 days from now
        deadline = (datetime.now() + timedelta(days=7)).isoformat()
        with open(SETTINGS_FILE, 'w') as f:
            json.dump({'deadline': deadline, 'is_locked': False}, f)
    
    if not os.path.exists(ASSIGNMENTS_FILE):
        with open(ASSIGNMENTS_FILE, 'w') as f:
            json.dump({}, f)

init_data_files()

# Helper functions
def load_json(filepath):
    with open(filepath, 'r') as f:
        return json.load(f)

def save_json(filepath, data):
    with open(filepath, 'w') as f:
        json.dump(data, f, indent=2)

def get_employees():
    return load_json(EMPLOYEES_FILE)

def get_preferences():
    return load_json(PREFERENCES_FILE)

def get_settings():
    return load_json(SETTINGS_FILE)

def get_assignments():
    return load_json(ASSIGNMENTS_FILE)

def format_deadline(iso_datetime_str):
    """Format ISO datetime to readable format: 'Nov. 27, 2025 3:24 a.m. ET'"""
    dt = datetime.fromisoformat(iso_datetime_str)
    month = dt.strftime('%b')
    day = dt.day
    year = dt.year
    hour = dt.hour
    minute = dt.minute
    
    # Convert to 12-hour format with am/pm
    if hour == 0:
        hour_12 = 12
        am_pm = 'a.m.'
    elif hour < 12:
        hour_12 = hour
        am_pm = 'a.m.'
    elif hour == 12:
        hour_12 = 12
        am_pm = 'p.m.'
    else:
        hour_12 = hour - 12
        am_pm = 'p.m.'
    
    return f"{month}. {day}, {year} {hour_12}:{minute:02d} {am_pm} ET"

def calculate_satisfaction_score(preferences, assigned_shift_id):
    """
    Calculate satisfaction score for a single assigned shift.
    Lower score = better (based on preference rank)
    """
    top_prefs = preferences.get('top_12', [])
    
    if assigned_shift_id in top_prefs:
        rank = top_prefs.index(assigned_shift_id) + 1
        return rank
    else:
        # Not in top preferences - assign high penalty score
        return 999

def has_same_weekend_conflict(employee_shifts, new_shift_id):
    """
    Check if assigning new_shift_id would create two shifts on same weekend.
    Returns True if there's a conflict.
    """
    new_shift = next(s for s in SHIFTS if s['id'] == new_shift_id)
    new_week = new_shift['week']
    
    # Check all employee's existing shifts
    for shift_id in employee_shifts:
        existing_shift = next(s for s in SHIFTS if s['id'] == shift_id)
        if existing_shift['week'] == new_week:
            return True
    
    return False

def has_consecutive_shift_conflict(employee_shifts, new_shift_id):
    """
    Check if assigning new_shift_id would create consecutive/overlapping shifts.
    This catches Sunday 8am-4pm + Sunday 3pm-10pm on same day.
    Returns True if there's a conflict.
    """
    new_shift = next(s for s in SHIFTS if s['id'] == new_shift_id)
    new_date = new_shift['date']
    new_day = new_shift['day']
    
    # Check all employee's existing shifts
    for shift_id in employee_shifts:
        existing_shift = next(s for s in SHIFTS if s['id'] == shift_id)
        
        # If same date and same day (Sunday has 2 shifts), it's consecutive/overlapping
        if existing_shift['date'] == new_date and existing_shift['day'] == new_day:
            return True
    
    return False

# Routes
@app.route('/')
def index():
    if 'username' in session:
        if session.get('is_manager'):
            return redirect(url_for('manager_dashboard'))
        else:
            return redirect(url_for('employee_dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        data = request.json
        username = data.get('username')
        password = data.get('password')
        
        employees = get_employees()
        
        if username in employees:
            if check_password_hash(employees[username]['password'], password):
                session['username'] = username
                session['is_manager'] = employees[username].get('is_manager', False)
                return jsonify({'success': True, 'is_manager': session['is_manager']})
        
        return jsonify({'success': False, 'message': 'Invalid credentials'}), 401
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/manager/dashboard')
def manager_dashboard():
    if not session.get('is_manager'):
        return redirect(url_for('login'))
    
    employees = get_employees()
    settings = get_settings()
    preferences = get_preferences()
    assignments = get_assignments()
    
    # Count employees who have submitted preferences (top 12 + bottom 6)
    submitted_count = sum(1 for emp, prefs in preferences.items() 
                         if prefs and len(prefs.get('top_12', [])) == 12 and len(prefs.get('bottom_6', [])) == 6)
    
    return render_template('manager_dashboard.html', 
                         employees=employees,
                         settings=settings,
                         submitted_count=submitted_count,
                         total_employees=len([e for e in employees.values() if not e.get('is_manager')]),
                         assignments=assignments,
                         preferences=preferences,
                         shifts=SHIFTS)

@app.route('/employee/dashboard')
def employee_dashboard():
    if 'username' not in session or session.get('is_manager'):
        return redirect(url_for('login'))
    
    settings = get_settings()
    preferences = get_preferences()
    assignments = get_assignments()
    username = session['username']
    
    user_prefs = preferences.get(username, {})
    user_assignments = assignments.get(username, [])
    
    # Check if deadline has passed
    deadline = datetime.fromisoformat(settings['deadline'])
    is_locked = settings.get('is_locked', False) or datetime.now() > deadline
    
    # Format deadline for display
    try:
        formatted_deadline = format_deadline(settings['deadline'])
    except Exception as e:
        # Fallback to original format if formatting fails
        formatted_deadline = settings['deadline']
    
    return render_template('employee_dashboard.html',
                         username=username,
                         shifts=SHIFTS,
                         preferences=user_prefs,
                         assignments=user_assignments,
                         deadline=formatted_deadline,
                         is_locked=is_locked)

@app.route('/api/employees', methods=['GET', 'POST', 'DELETE'])
def manage_employees():
    if not session.get('is_manager'):
        return jsonify({'error': 'Unauthorized'}), 403
    
    employees = get_employees()
    
    if request.method == 'POST':
        data = request.json
        username = data.get('username')
        password = data.get('password')
        name = data.get('name')
        
        if username in employees:
            return jsonify({'error': 'Employee already exists'}), 400
        
        employees[username] = {
            'name': name,
            'password': generate_password_hash(password),
            'is_manager': False
        }
        save_json(EMPLOYEES_FILE, employees)
        return jsonify({'success': True})
    
    elif request.method == 'DELETE':
        data = request.json
        username = data.get('username')
        
        if username in employees:
            del employees[username]
            save_json(EMPLOYEES_FILE, employees)
            
            # Also remove their preferences
            preferences = get_preferences()
            if username in preferences:
                del preferences[username]
                save_json(PREFERENCES_FILE, preferences)
            
            return jsonify({'success': True})
        
        return jsonify({'error': 'Employee not found'}), 404
    
    # GET
    return jsonify(employees)

@app.route('/api/preferences', methods=['GET', 'POST'])
def manage_preferences():
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 403
    
    username = session['username']
    preferences = get_preferences()
    settings = get_settings()
    
    # Check if locked
    deadline = datetime.fromisoformat(settings['deadline'])
    is_locked = settings.get('is_locked', False) or datetime.now() > deadline
    
    if request.method == 'POST':
        if is_locked and not session.get('is_manager'):
            return jsonify({'error': 'Preferences are locked'}), 403
        
        data = request.json
        
        # Validate data structure
        if 'top_12' not in data or 'bottom_6' not in data or 'shift_type_pref' not in data:
            return jsonify({'error': 'Invalid preference format'}), 400
        
        if len(data['top_12']) != 12:
            return jsonify({'error': 'Must select exactly 12 top preferences'}), 400
        
        if len(data['bottom_6']) != 6:
            return jsonify({'error': 'Must select exactly 6 least wanted shifts'}), 400
        
        preferences[username] = {
            'top_12': data['top_12'],
            'bottom_6': data['bottom_6'],
            'shift_type_pref': data['shift_type_pref']
        }
        save_json(PREFERENCES_FILE, preferences)
        return jsonify({'success': True})
    
    # GET
    if session.get('is_manager'):
        return jsonify(preferences)
    else:
        return jsonify({username: preferences.get(username, {})})

@app.route('/api/settings', methods=['GET', 'POST'])
def manage_settings():
    if not session.get('is_manager'):
        return jsonify({'error': 'Unauthorized'}), 403
    
    settings = get_settings()
    
    if request.method == 'POST':
        data = request.json
        
        if 'deadline' in data:
            settings['deadline'] = data['deadline']
        
        if 'is_locked' in data:
            settings['is_locked'] = data['is_locked']
        
        save_json(SETTINGS_FILE, settings)
        return jsonify({'success': True})
    
    return jsonify(settings)

@app.route('/api/allocate', methods=['POST'])
def allocate_shifts():
    if not session.get('is_manager'):
        return jsonify({'error': 'Unauthorized'}), 403
    
    preferences = get_preferences()
    employees_data = get_employees()
    
    # Get list of non-manager employees
    employee_list = [user for user, emp in employees_data.items() if not emp.get('is_manager')]
    
    # Separate employees into two groups:
    # 1. Those with complete preferences (top 12 + bottom 6)
    # 2. Those without complete preferences (will be randomly assigned)
    employees_with_prefs = []
    employees_without_prefs = []
    
    for emp in employee_list:
        if emp in preferences:
            prefs = preferences[emp]
            if len(prefs.get('top_12', [])) == 12 and len(prefs.get('bottom_6', [])) == 6:
                employees_with_prefs.append(emp)
            else:
                employees_without_prefs.append(emp)
        else:
            employees_without_prefs.append(emp)
    
    # Initialize assignments
    assignments = {emp: [] for emp in employee_list}
    shift_assignments = {shift['id']: [] for shift in SHIFTS}
    warnings = []  # Track employees who were randomly assigned
    
    # Set random seed for reproducibility
    random.seed(42)
    
    # PHASE 1: Preference-based allocation for employees with complete preferences
    print("\n=== PHASE 1: Preference-Based Allocation ===")
    print(f"Processing {len(employees_with_prefs)} employees with complete preferences\n")
    
    shuffled_employees = employees_with_prefs.copy()
    random.shuffle(shuffled_employees)
    
    for emp in shuffled_employees:
        prefs = preferences[emp]
        top_12 = prefs['top_12']
        bottom_6 = prefs['bottom_6']
        
        # Try to assign from top 12 preferences
        assigned = False
        for shift_id in top_12:
            # Skip if shift is full
            shift = next(s for s in SHIFTS if s['id'] == shift_id)
            if len(shift_assignments[shift_id]) >= shift['slots']:
                continue
            
            # Skip if would create same-weekend conflict for second shift
            if has_same_weekend_conflict(assignments[emp], shift_id):
                continue
            
            # Skip if would create consecutive shift conflict
            if has_consecutive_shift_conflict(assignments[emp], shift_id):
                continue
            
            # Assign shift
            assignments[emp].append(shift_id)
            shift_assignments[shift_id].append(emp)
            assigned = True
            rank = top_12.index(shift_id) + 1
            print(f"✓ {emp:15} → Shift {shift_id:2} (preference #{rank})")
            break
        
        # If couldn't assign from top 12, try non-bottom-6 shifts
        if not assigned:
            shift_type_pref = prefs.get('shift_type_pref', {})
            # Sort shift types by preference (1=best, 3=worst)
            sorted_types = sorted(shift_type_pref.items(), key=lambda x: x[1])
            
            for shift_type, _ in sorted_types:
                for shift in SHIFTS:
                    shift_id = shift['id']
                    
                    # Skip if in bottom 6
                    if shift_id in bottom_6:
                        continue
                    
                    # Skip if already assigned this shift
                    if shift_id in top_12:  # Already tried these
                        continue
                    
                    # Check if matches shift type
                    shift_matches = False
                    if shift_type == 'saturday' and shift['day'] == 'Saturday':
                        shift_matches = True
                    elif shift_type == 'sunday_morning' and shift['day'] == 'Sunday' and '8:00 AM' in shift['time']:
                        shift_matches = True
                    elif shift_type == 'sunday_evening' and shift['day'] == 'Sunday' and '3:00 PM' in shift['time']:
                        shift_matches = True
                    
                    if not shift_matches:
                        continue
                    
                    # Skip if shift is full
                    if len(shift_assignments[shift_id]) >= shift['slots']:
                        continue
                    
                    # Skip if would create same-weekend conflict
                    if has_same_weekend_conflict(assignments[emp], shift_id):
                        continue
                    
                    # Skip if would create consecutive shift conflict
                    if has_consecutive_shift_conflict(assignments[emp], shift_id):
                        continue
                    
                    # Assign shift
                    assignments[emp].append(shift_id)
                    shift_assignments[shift_id].append(emp)
                    assigned = True
                    print(f"⚠ {emp:15} → Shift {shift_id:2} (backup assignment, not in top 12)")
                    break
                
                if assigned:
                    break
        
        if not assigned:
            print(f"✗ {emp:15} → Could not assign first shift")
    
    # PHASE 2: Second shift allocation for employees with preferences (sorted by satisfaction from Phase 1)
    print("\n=== PHASE 2: Second Shift Allocation (Preference-Based) ===")
    
    # Calculate satisfaction scores from Phase 1
    employee_satisfaction = []
    for emp in employees_with_prefs:
        if len(assignments[emp]) > 0:
            score = calculate_satisfaction_score(preferences[emp], assignments[emp][0])
            employee_satisfaction.append((emp, score))
        else:
            employee_satisfaction.append((emp, 9999))  # No first shift = highest priority
    
    # Sort by satisfaction (worst first), then randomize within same score
    employee_satisfaction.sort(key=lambda x: (x[1], random.random()))
    
    for emp, phase1_score in employee_satisfaction:
        # Skip if already has 2 shifts
        if len(assignments[emp]) >= 2:
            continue
        
        prefs = preferences[emp]
        top_12 = prefs['top_12']
        bottom_6 = prefs['bottom_6']
        
        # Try to assign from top 12 preferences
        assigned = False
        for shift_id in top_12:
            # Skip if already assigned this shift
            if shift_id in assignments[emp]:
                continue
            
            # Skip if shift is full
            shift = next(s for s in SHIFTS if s['id'] == shift_id)
            if len(shift_assignments[shift_id]) >= shift['slots']:
                continue
            
            # Skip if would create same-weekend conflict
            if has_same_weekend_conflict(assignments[emp], shift_id):
                continue
            
            # Skip if would create consecutive shift conflict
            if has_consecutive_shift_conflict(assignments[emp], shift_id):
                continue
            
            # Assign shift
            assignments[emp].append(shift_id)
            shift_assignments[shift_id].append(emp)
            assigned = True
            rank = top_12.index(shift_id) + 1
            print(f"✓ {emp:15} → Shift {shift_id:2} (preference #{rank}, phase 1 score was {phase1_score})")
            break
        
        # If couldn't assign from top 12, try non-bottom-6 shifts
        if not assigned:
            shift_type_pref = prefs.get('shift_type_pref', {})
            sorted_types = sorted(shift_type_pref.items(), key=lambda x: x[1])
            
            for shift_type, _ in sorted_types:
                for shift in SHIFTS:
                    shift_id = shift['id']
                    
                    # Skip if already assigned
                    if shift_id in assignments[emp]:
                        continue
                    
                    # Skip if in bottom 6
                    if shift_id in bottom_6:
                        continue
                    
                    # Check if matches shift type
                    shift_matches = False
                    if shift_type == 'saturday' and shift['day'] == 'Saturday':
                        shift_matches = True
                    elif shift_type == 'sunday_morning' and shift['day'] == 'Sunday' and '8:00 AM' in shift['time']:
                        shift_matches = True
                    elif shift_type == 'sunday_evening' and shift['day'] == 'Sunday' and '3:00 PM' in shift['time']:
                        shift_matches = True
                    
                    if not shift_matches:
                        continue
                    
                    # Skip if shift is full
                    if len(shift_assignments[shift_id]) >= shift['slots']:
                        continue
                    
                    # Skip if would create same-weekend conflict
                    if has_same_weekend_conflict(assignments[emp], shift_id):
                        continue
                    
                    # Assign shift
                    assignments[emp].append(shift_id)
                    shift_assignments[shift_id].append(emp)
                    assigned = True
                    print(f"⚠ {emp:15} → Shift {shift_id:2} (backup assignment, not in top 12)")
                    break
                
                if assigned:
                    break
        
        if not assigned:
            print(f"✗ {emp:15} → Could not assign second shift")
    
    # PHASE 3: Random assignment for employees without complete preferences
    if employees_without_prefs:
        print("\n=== PHASE 3: Random Assignment for Laggards ===")
        print(f"Processing {len(employees_without_prefs)} employees without complete preferences\n")
        
        for emp in employees_without_prefs:
            warnings.append(f"{emp} was randomly assigned (no preferences submitted)")
            
            # Get all available shifts (not full, not creating weekend conflicts)
            available_shifts = []
            for shift in SHIFTS:
                shift_id = shift['id']
                
                # Skip if shift is full
                if len(shift_assignments[shift_id]) >= shift['slots']:
                    continue
                
                # Skip if would create same-weekend conflict
                if has_same_weekend_conflict(assignments[emp], shift_id):
                    continue
                
                # Skip if would create consecutive shift conflict
                if has_consecutive_shift_conflict(assignments[emp], shift_id):
                    continue
                
                available_shifts.append(shift_id)
            
            # Randomly assign 2 shifts from available shifts
            if len(available_shifts) >= 2:
                selected_shifts = random.sample(available_shifts, 2)
                for shift_id in selected_shifts:
                    assignments[emp].append(shift_id)
                    shift_assignments[shift_id].append(emp)
                    print(f"⚠ {emp:15} → Shift {shift_id:2} (random assignment)")
            else:
                print(f"✗ {emp:15} → Not enough available shifts for random assignment")
                warnings.append(f"{emp} could not be fully assigned - insufficient available shifts")
    
    # Save assignments
    save_json(ASSIGNMENTS_FILE, assignments)
    
    # Lock preferences
    settings = get_settings()
    settings['is_locked'] = True
    save_json(SETTINGS_FILE, settings)
    
    return jsonify({
        'success': True,
        'assignments': assignments,
        'shift_assignments': shift_assignments,
        'warnings': warnings
    })

@app.route('/api/backup')
def backup_data():
    """Download all data files as JSON for backup"""
    if not session.get('is_manager'):
        return jsonify({'error': 'Unauthorized'}), 403
    
    backup_data = {
        'employees': get_employees(),
        'preferences': get_preferences(),
        'settings': get_settings(),
        'assignments': get_assignments(),
        'timestamp': datetime.now().isoformat()
    }
    
    from io import BytesIO
    output = BytesIO()
    output.write(json.dumps(backup_data, indent=2).encode('utf-8'))
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/json',
        as_attachment=True,
        download_name=f'backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'
    )

@app.route('/api/populate-test-data', methods=['POST'])
def populate_test_data():
    """Populate random preferences for all employees (TESTING ONLY)"""
    if not session.get('is_manager'):
        return jsonify({'error': 'Unauthorized'}), 403
    
    employees = get_employees()
    preferences = {}
    
    # All 60 shift IDs
    all_shifts = list(range(60))
    
    # Generate random preferences for each non-manager employee
    for username, emp_data in employees.items():
        if emp_data.get('is_manager'):
            continue
        
        # Shuffle all shifts
        shuffled = all_shifts.copy()
        random.shuffle(shuffled)
        
        # Top 12 are first 12 from shuffled list
        top_12 = shuffled[:12]
        
        # Bottom 6 are next 6 from shuffled list
        bottom_6 = shuffled[12:18]
        
        # Random shift type preferences (1, 2, 3)
        shift_types = [1, 2, 3]
        random.shuffle(shift_types)
        
        preferences[username] = {
            'top_12': top_12,
            'bottom_6': bottom_6,
            'shift_type_pref': {
                'saturday': str(shift_types[0]),
                'sunday_morning': str(shift_types[1]),
                'sunday_evening': str(shift_types[2])
            }
        }
    
    # Save preferences
    save_json(PREFERENCES_FILE, preferences)
    
    return jsonify({
        'success': True,
        'message': f'Populated random preferences for {len(preferences)} employees'
    })

@app.route('/api/export-excel')
def export_excel():
    if not session.get('is_manager'):
        return jsonify({'error': 'Unauthorized'}), 403
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO
        
        assignments = get_assignments()
        employees = get_employees()
        preferences = get_preferences()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Weekend Schedule"
        
        # Title
        ws['A1'] = 'Weekend Trunk Shift Schedule - Dec 2025 - Apr 2026'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:H1')
        
        # Headers
        headers = ['Date', 'Day', 'Time', 'Assigned Employee', 'Preference Rank', 'Status', 'Week', 'Notes']
        header_row = 3
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data
        row = header_row + 1
        for shift in SHIFTS:
            shift_id = shift['id']
            assigned = []
            
            for emp, emp_shifts in assignments.items():
                if shift_id in emp_shifts:
                    assigned.append(emp)
            
            # Date
            ws.cell(row=row, column=1).value = shift['date']
            
            # Day
            ws.cell(row=row, column=2).value = shift['day']
            
            # Time
            ws.cell(row=row, column=3).value = shift['time']
            
            # Assigned employee
            if assigned:
                emp = assigned[0]
                emp_name = employees[emp]['name']
                ws.cell(row=row, column=4).value = emp_name
                
                # Get preference rank
                if emp in preferences:
                    prefs = preferences[emp]
                    if shift_id in prefs.get('top_12', []):
                        rank = prefs['top_12'].index(shift_id) + 1
                        ws.cell(row=row, column=5).value = f"#{rank}"
                    elif shift_id in prefs.get('bottom_6', []):
                        ws.cell(row=row, column=5).value = "Bottom 6"
                        ws.cell(row=row, column=5).font = Font(color="FF0000")
                    else:
                        ws.cell(row=row, column=5).value = "N/A"
                else:
                    ws.cell(row=row, column=5).value = "N/A"
            else:
                ws.cell(row=row, column=4).value = "VACANT"
                ws.cell(row=row, column=4).font = Font(color="FF0000", bold=True)
            
            # Status
            if assigned:
                ws.cell(row=row, column=6).value = "FILLED"
                ws.cell(row=row, column=6).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                ws.cell(row=row, column=6).value = "VACANT"
                ws.cell(row=row, column=6).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            # Week
            ws.cell(row=row, column=7).value = shift['week']
            
            row += 1
        
        # Employee summary section
        row += 2
        ws.cell(row=row, column=1).value = "Employee Summary"
        ws.cell(row=row, column=1).font = Font(size=14, bold=True)
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        summary_headers = ['Employee', 'Shifts Assigned', 'Shift Details', 'Status']
        for col, header in enumerate(summary_headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        row += 1
        for emp, emp_data in employees.items():
            if emp_data.get('is_manager'):
                continue
            
            ws.cell(row=row, column=1).value = emp_data['name']
            
            emp_shifts = assignments.get(emp, [])
            ws.cell(row=row, column=2).value = len(emp_shifts)
            
            # Shift details
            shift_details = []
            for shift_id in emp_shifts:
                shift = next(s for s in SHIFTS if s['id'] == shift_id)
                shift_details.append(f"{shift['date']} {shift['day']} {shift['time']}")
            ws.cell(row=row, column=3).value = "; ".join(shift_details) if shift_details else "None"
            
            # Status
            if len(emp_shifts) == 2:
                ws.cell(row=row, column=4).value = "Complete"
                ws.cell(row=row, column=4).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                ws.cell(row=row, column=4).value = f"Incomplete ({len(emp_shifts)}/2)"
                ws.cell(row=row, column=4).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 30
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'weekend_shift_schedule_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/change-password', methods=['POST'])
def change_password():
    """Allow users to change their password"""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 403
    
    username = session['username']
    data = request.json
    
    current_password = data.get('current_password')
    new_password = data.get('new_password')
    
    if not current_password or not new_password:
        return jsonify({'error': 'Current and new password required'}), 400
    
    if len(new_password) < 6:
        return jsonify({'error': 'New password must be at least 6 characters'}), 400
    
    employees = get_employees()
    
    # Verify current password
    if username not in employees:
        return jsonify({'error': 'User not found'}), 404
    
    if not check_password_hash(employees[username]['password'], current_password):
        return jsonify({'error': 'Current password is incorrect'}), 401
    
    # Update password
    employees[username]['password'] = generate_password_hash(new_password)
    save_json(EMPLOYEES_FILE, employees)
    
    return jsonify({'success': True, 'message': 'Password changed successfully'})

# Trunk writer credentials data (embedded to avoid CSV file dependency)
TRUNK_WRITER_CREDENTIALS = [
    {"name": "Ali, Idrees", "username": "idrees.ali", "password": "2lf92k"},
    {"name": "Allen, Jonathan", "username": "jonathan.allen", "password": "Rml6iy"},
    {"name": "Alper, Alexandra", "username": "alexandra.alper", "password": "2xgrpb"},
    {"name": "Ax, Joseph A.", "username": "joseph.ax", "password": "svwBap"},
    {"name": "Brooks, Brad", "username": "brad.brooks", "password": "PXU389"},
    {"name": "Brunnstrom, David R.", "username": "david.brunnstrom", "password": "vHkw82"},
    {"name": "Coster, Helen A.", "username": "helen.coster", "password": "5KPZyQ"},
    {"name": "Cowan, Richard J.", "username": "richard.cowan", "password": "NztlMK"},
    {"name": "Erickson, Bo", "username": "bo.erickson", "password": "c5zC2l"},
    {"name": "Gorman, Steve J.", "username": "steve.gorman", "password": "EvxNGj"},
    {"name": "Goudsward, Andrew", "username": "andrew.goudsward", "password": "BOQXUQ"},
    {"name": "Harte, Julia", "username": "julia.harte", "password": "dHSGh4"},
    {"name": "Heath, Brad", "username": "brad.heath", "password": "014mtj"},
    {"name": "Heavey, Susan", "username": "sheavey", "password": "Y5QaRO"},
    {"name": "Hesson, Ted", "username": "ted.hesson", "password": "x2vGuM"},
    {"name": "Kruzel, John", "username": "john.kruzel", "password": "YH3hq7"},
    {"name": "Lawder, David", "username": "david.lawder", "password": "p5IBz2"},
    {"name": "Layne, Nathan", "username": "nathan.layne", "password": "m2D50J"},
    {"name": "Lewis, Simon", "username": "simon.lewis", "password": "CBm7wr"},
    {"name": "Martina, Michael", "username": "michael.martina", "password": "HZPo9k"},
    {"name": "Morgan, David G.", "username": "david.morgan", "password": "qVO46H"},
    {"name": "Oliphant, Jim", "username": "james.oliphant", "password": "usJlEx"},
    {"name": "Reid, Timothy J.", "username": "tim.reid", "password": "KBYjFi"},
    {"name": "Satter, Raphael", "username": "raphael.satter", "password": "3d4NLr"},
    {"name": "Spetalnick, Matt S.", "username": "matt.spetalnick", "password": "DlBsRU"},
    {"name": "Stewart, Phillip", "username": "phillip.stewart", "password": "iY0ChY"},
    {"name": "Sullivan, Andy", "username": "andy.sullivan", "password": "3J0zdD"},
    {"name": "Tanfani, Joseph", "username": "joseph.tanfani", "password": "S0Eu19"},
    {"name": "Trotta, Daniel", "username": "daniel.trotta", "password": "43pdyN"},
    {"name": "Zengerle, Patricia A.", "username": "patricia.zengerle", "password": "US1VPF"},
]

@app.route('/api/reload-employees-from-csv', methods=['POST'])
def reload_employees_from_csv():
    """Reload employee accounts from embedded credentials data (ADMIN ONLY)"""
    if not session.get('is_manager'):
        return jsonify({'error': 'Unauthorized'}), 403
    
    try:
        employees = {}
        
        # Add admin account
        employees['admin'] = {
            'name': 'Admin',
            'is_manager': True,
            'password': generate_password_hash('admin123')
        }
        
        # Add all trunk writers from embedded data
        for writer in TRUNK_WRITER_CREDENTIALS:
            username = writer['username']
            name = writer['name']
            password = writer['password']
            
            employees[username] = {
                'name': name,
                'is_manager': False,
                'password': generate_password_hash(password)
            }
        
        # Save to employees.json
        save_json(EMPLOYEES_FILE, employees)
        
        return jsonify({
            'success': True,
            'message': f'Successfully reloaded {len(employees) - 1} trunk writer accounts',
            'total_accounts': len(employees)
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/reset-data', methods=['POST'])
def reset_data():
    """Reset preferences and assignments (ADMIN ONLY - for testing)"""
    if not session.get('is_manager'):
        return jsonify({'error': 'Unauthorized'}), 403
    
    try:
        # Clear preferences
        save_json(PREFERENCES_FILE, {})
        
        # Clear assignments
        save_json(ASSIGNMENTS_FILE, {})
        
        # Unlock preferences
        settings = get_settings()
        settings['is_locked'] = False
        save_json(SETTINGS_FILE, settings)
        
        return jsonify({
            'success': True,
            'message': 'All preferences and assignments cleared. System unlocked.'
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, port=5000)
